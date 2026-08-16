# -*- coding: utf-8 -*-
"""Excel'deki geçmiş DİİB kayıtlarını (2024/D1-03798 dönemi) sisteme aktarır.

Kaynaklar:
- İTHALAT sayfası      → ithalat + kalemler (renk dökümü satırları dahil)
- İHR.FAT. SARF sayfası → ihracat fatura kalemleri (satır kodu + kg — en güvenilir kaynak)
- İHRACAT sayfası      → beyanname no, müşteri, ülke, tutar, kapanış tarihi (fatura no ile birleştirilir)

Çalıştırma:  python -m backend.import_history "<excel yolu>"
Tekrar çalıştırılırsa 'excel-aktarim' kaynaklı eski kayıtları silip yeniden yükler (idempotent).
"""
import os
import re
import sys

import openpyxl

from . import db as dbm

EXCEL_DEFAULT = os.path.join(os.path.dirname(dbm.BASE_DIR), "DİİB-4 İTHALAT -İHRACAT SARF TABLOLARI (Otomatik kaydedildi).xlsx")


def _date(v):
    if v is None:
        return ""
    s = str(v)
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})", s)
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m else ""


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def import_ithalat(conn, wb, belge_id, ham_ids):
    ws = wb["İTHALAT"]
    rows = list(ws.iter_rows(min_row=8, max_row=40, min_col=1, max_col=30, values_only=True))
    count = 0
    current = None  # (ithalat_id, doviz)

    for r in rows:
        (sira, gumruk, _c, tescil, tarih, _f, alici, _h, _i, gtip, satici, mense,
         madde, kg, _o, bf, fk_doviz, doviz_cinsi, fk_tl, ist_kiymet, para, *_rest) = r[:21]
        kur = r[25]

        if sira is not None and (madde or tescil):
            # ana beyanname satırı
            doviz = str(para or "USD").replace("EURO", "EUR")
            cur = conn.execute(
                """INSERT INTO ithalat (belge_id, beyanname_no, tarih, gumruk, satici, mense, doviz,
                       tutar, kur, notlar, kaynak)
                   VALUES (?,?,?,?,?,?,?,?,?,?, 'excel-aktarim')""",
                (belge_id, str(tescil or "").strip(), _date(tarih), str(gumruk or "").strip(),
                 str(satici or "").strip(), str(mense or "").strip(), doviz,
                 _num(ist_kiymet) or _num(fk_doviz), _num(kur),
                 "" if tescil else "Beyanname bilgisi kaynak dosyada eksik"))
            iid = cur.lastrowid
            count += 1
            canon = dbm.canonical_hammadde(str(madde or ""))
            renkli_gelecek = str(madde or "").strip().upper() in ("PİGMENT", "PIGMENT")
            if canon in ham_ids and _num(kg) > 0 and not renkli_gelecek:
                conn.execute(
                    "INSERT INTO ithalat_kalem (ithalat_id, hammadde_id, aciklama, miktar_kg, birim_fiyat, tutar) VALUES (?,?,?,?,?,?)",
                    (iid, ham_ids[canon], str(madde).strip(), _num(kg), _num(bf),
                     _num(ist_kiymet) or round(_num(kg) * _num(bf), 2)))
            current = (iid, canon, _num(kg), _num(bf), _num(ist_kiymet), renkli_gelecek, [])
        elif current and madde and _num(bf) and current[5]:
            # pigment renk dökümü satırı: "RED 53.1   3000 Kg."
            m = re.search(r"([\d.,]+)\s*KG", str(madde).upper().replace(" ", " "))
            qty = 0.0
            if m:
                qty = _num(m.group(1).replace(".", "").replace(",", "."))
            renk = re.sub(r"[\d.,]+\s*KG\.?", "", str(madde), flags=re.IGNORECASE).strip()
            if qty > 0:
                conn.execute(
                    "INSERT INTO ithalat_kalem (ithalat_id, hammadde_id, aciklama, miktar_kg, birim_fiyat, tutar) VALUES (?,?,?,?,?,?)",
                    (current[0], ham_ids["PİGMENT"], f"PİGMENT {renk}", qty, _num(bf),
                     _num(ist_kiymet) or round(qty * _num(bf), 2)))

    # renk dökümü verilmemiş pigment ana satırları için toplam kalem oluştu mu kontrolü
    for row in conn.execute(
        """SELECT i.id, i.beyanname_no FROM ithalat i
           WHERE i.kaynak='excel-aktarim' AND i.belge_id=?
             AND NOT EXISTS (SELECT 1 FROM ithalat_kalem k WHERE k.ithalat_id=i.id)""", (belge_id,)).fetchall():
        conn.execute("UPDATE ithalat SET notlar = notlar || ' | Kalem dökümü Excel''den okunamadı, elle tamamlayın' WHERE id=?", (row["id"],))
    return count


def import_ihracat(conn, wb, belge_id, mamul_by_kod):
    # 1) İHRACAT sayfası: beyanname başlıkları (fatura no → başlık bilgisi)
    ws = wb["İHRACAT"]
    headers = {}
    for r in ws.iter_rows(min_row=4, max_row=60, min_col=1, max_col=18, values_only=True):
        sira, gumruk, beyanname, btarih, fatura, ftarih, musteri, ulke, kg, eur, usd, tl, *_r = r[:12]
        kapanis = r[17]
        if sira is None or not fatura:
            continue
        fno = str(fatura).strip().upper().replace("İ", "I")
        headers[fno] = {
            "beyanname_no": str(beyanname or "").strip(),
            "tarih": _date(ftarih) or _date(btarih),
            "gumruk": str(gumruk or "").strip(),
            "musteri": str(musteri or "").strip(),
            "ulke": str(ulke or "").strip(),
            "doviz": "EUR" if _num(eur) else "USD",
            "tutar": _num(eur) or _num(usd),
            "kapanis": _date(kapanis) if kapanis and str(kapanis).startswith("20") else ("AÇIK" if str(kapanis).strip() == "AÇIK" else _date(kapanis)),
        }

    # 2) İHR.FAT. SARF: fatura kalemleri (satır kodu + kg)
    ws = wb["İHR.FAT. SARF"]
    count = 0
    current_id = None
    current_fno = None
    for r in ws.iter_rows(min_row=2, max_row=600, min_col=1, max_col=12, values_only=True):
        sira, fatura, tarih, _t2, musteri, ulke, gtip, satir_kodu, _fsn, urun, cins, kg = r
        urun_s = str(urun or "").strip()
        if urun_s.upper().startswith(("AĞUSTOS", "EYLÜL", "EKİM", "KASIM", "ARALIK", "OCAK", "ŞUBAT",
                                      "MART", "NİSAN", "MAYIS", "HAZİRAN", "TEMMUZ", "KÜMÜLATİF")) or "TOPLAM" in urun_s.upper():
            continue
        if fatura:  # yeni fatura başlıyor
            fno = str(fatura).strip().upper().replace("İ", "I")
            h = headers.get(fno, {})
            cur = conn.execute(
                """INSERT INTO ihracat (belge_id, fatura_no, beyanname_no, tarih, musteri, ulke, gumruk,
                       doviz, tutar, kapanis_tarihi, kaynak)
                   VALUES (?,?,?,?,?,?,?,?,?,?, 'excel-aktarim')""",
                (belge_id, str(fatura).strip(), h.get("beyanname_no", ""),
                 h.get("tarih") or _date(tarih), str(musteri or h.get("musteri", "")).strip(),
                 str(ulke or h.get("ulke", "")).strip(), h.get("gumruk", ""),
                 h.get("doviz", "USD"), h.get("tutar", 0), h.get("kapanis", "")))
            current_id = cur.lastrowid
            current_fno = fno
            count += 1
        if current_id is None or not urun_s or _num(kg) <= 0:
            continue
        kod = str(satir_kodu or "").strip()
        mamul = mamul_by_kod.get(kod)
        if not mamul:
            # revize ile eklenen ek satır kodları aynı ürün grubuna eşlenir
            g = kod.split(".")[-1] if "." in kod else ""
            g = {"016": "006", "017": "007", "018": "008"}.get(g, g)
            mamul = next((m for k2, m in mamul_by_kod.items() if k2.endswith("." + g)), None) if g else None
        if not mamul:
            continue
        conn.execute(
            "INSERT INTO ihracat_kalem (ihracat_id, mamul_id, urun_adi, miktar_kg) VALUES (?,?,?,?)",
            (current_id, mamul["id"], urun_s[:80], _num(kg)))
    return count


def run(excel_path: str = EXCEL_DEFAULT):
    dbm.init_db()
    conn = dbm.get_conn()
    belge = conn.execute("SELECT * FROM belge LIMIT 1").fetchone()
    belge_id = belge["id"]
    firma_id = belge["firma_id"]

    # idempotent: eski aktarımı temizle
    conn.execute("DELETE FROM ithalat WHERE kaynak='excel-aktarim' AND belge_id=?", (belge_id,))
    conn.execute("DELETE FROM ihracat WHERE kaynak='excel-aktarim' AND belge_id=?", (belge_id,))

    ham_ids = {r["ad"]: r["id"] for r in conn.execute("SELECT id, ad FROM hammadde WHERE firma_id=?", (firma_id,))}
    mamul_by_kod = {r["satir_kodu"]: dict(r) for r in conn.execute("SELECT * FROM mamul WHERE belge_id=?", (belge_id,))}

    wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
    n_ith = import_ithalat(conn, wb, belge_id, ham_ids)
    n_ihr = import_ihracat(conn, wb, belge_id, mamul_by_kod)
    conn.commit()

    kalem_ith = conn.execute("SELECT COUNT(*) c FROM ithalat_kalem k JOIN ithalat i ON i.id=k.ithalat_id WHERE i.kaynak='excel-aktarim'").fetchone()["c"]
    kalem_ihr = conn.execute("SELECT COUNT(*) c FROM ihracat_kalem k JOIN ihracat i ON i.id=k.ihracat_id WHERE i.kaynak='excel-aktarim'").fetchone()["c"]
    toplam_kg = conn.execute("SELECT COALESCE(SUM(k.miktar_kg),0) s FROM ihracat_kalem k JOIN ihracat i ON i.id=k.ihracat_id WHERE i.kaynak='excel-aktarim'").fetchone()["s"]
    conn.close()
    return {"ithalat": n_ith, "ithalat_kalem": kalem_ith, "ihracat": n_ihr,
            "ihracat_kalem": kalem_ihr, "ihracat_toplam_kg": toplam_kg}


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else EXCEL_DEFAULT
    print(run(path))
