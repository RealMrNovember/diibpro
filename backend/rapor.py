# -*- coding: utf-8 -*-
"""Resmî rapor çıktıları — tek veri yapısından üç format: XLSX · PDF · HTML (görüntüleme).

GET /api/rapor/{kdv|tev|kapatma}?format=xlsx|pdf|html   (varsayılan xlsx)

Tüm çıktılar "gümrük müşaviri / YMM onayına hazır TASLAK" niteliğindedir.
"""
import html as html_mod
import io
import os
from datetime import date

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import HTMLResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from . import auth
from . import db as dbm

router = APIRouter(prefix="/api/rapor")

TEV_ULKELERI = [
    "ALMANYA", "GERMANY", "FRANSA", "FRANCE", "İTALYA", "ITALY", "İSPANYA", "SPAIN",
    "HOLLANDA", "NETHERLANDS", "BELÇİKA", "BELGIUM", "POLONYA", "POLAND", "YUNANİSTAN",
    "GREECE", "PORTEKİZ", "PORTUGAL", "AVUSTURYA", "AUSTRIA", "İSVEÇ", "SWEDEN",
    "DANİMARKA", "DENMARK", "FİNLANDİYA", "FINLAND", "İRLANDA", "IRELAND", "ROMANYA",
    "ROMANIA", "BULGARİSTAN", "BULGARIA", "MACARİSTAN", "HUNGARY", "ÇEKYA", "CZECH",
    "SLOVAKYA", "SLOVAKIA", "SLOVENYA", "SLOVENIA", "HIRVATİSTAN", "CROATIA",
    "LİTVANYA", "LITHUANIA", "LETONYA", "LATVIA", "ESTONYA", "ESTONIA", "KIBRIS",
    "CYPRUS", "MALTA", "LÜKSEMBURG", "LUXEMBOURG",
    "ARNAVUTLUK", "ALBANIA", "ALBANİA", "GÜRCİSTAN", "GEORGIA", "GEORGİA",
    "MAKEDONYA", "MACEDONIA", "BOSNA", "BOSNIA", "SIRBİSTAN", "SERBIA",
    "KARADAĞ", "MONTENEGRO", "KOSOVA", "KOSOVO", "MOLDOVA", "İSRAİL", "ISRAEL",
    "MISIR", "EGYPT", "TUNUS", "TUNISIA", "FAS", "MOROCCO", "ŞİLİ", "CHILE",
    "GÜNEY KORE", "KOREA", "MALEZYA", "MALAYSIA", "SİNGAPUR", "SINGAPORE",
    "İNGİLTERE", "UNITED KINGDOM", "UK", "İSVİÇRE", "SWITZERLAND", "NORVEÇ", "NORWAY",
]
TEV_ORANI = 0.065
UCUNCU_ULKE_MENSE = ["ÇİN", "CHINA", "HİNDİSTAN", "INDIA"]


# ================================================================ VERİ YAPISI
# rapor = {"baslik", "altbilgi", "dosya", "sayfalar": [
#           {"ad", "basliklar":[...], "genislikler":[...], "num_cols":{...}, "satirlar":[[...]], "not": str} ]}

def _ctx(user):
    conn = dbm.get_conn()
    belge = dict(conn.execute("SELECT * FROM belge WHERE firma_id=? ORDER BY id DESC LIMIT 1",
                              (user["firma_id"],)).fetchone())
    firma = dict(conn.execute("SELECT * FROM firma WHERE id=?", (user["firma_id"],)).fetchone())
    return conn, belge, firma


def _recete_map(conn, bid):
    recete = {}
    for r in conn.execute(
        """SELECT r.mamul_id, r.hammadde_id, r.katsayi FROM recete r
           JOIN mamul m ON m.id=r.mamul_id WHERE m.belge_id=?""", (bid,)):
        recete.setdefault(r["mamul_id"], {})[r["hammadde_id"]] = r["katsayi"]
    return recete


def build_kdv(user):
    conn, belge, firma = _ctx(user)
    bid = belge["id"]
    ham = [dict(r) for r in conn.execute(
        """SELECT DISTINCT h.id, h.ad FROM ithalat_kalem k
           JOIN ithalat i ON i.id=k.ithalat_id JOIN hammadde h ON h.id=k.hammadde_id
           WHERE i.belge_id=? ORDER BY h.ad""", (bid,))]
    son_fiyat = {}
    for h in ham:
        r = conn.execute(
            """SELECT k.birim_fiyat, i.kur FROM ithalat_kalem k JOIN ithalat i ON i.id=k.ithalat_id
               WHERE i.belge_id=? AND k.hammadde_id=? AND k.birim_fiyat > 0
               ORDER BY i.tarih DESC LIMIT 1""", (bid, h["id"])).fetchone()
        son_fiyat[h["id"]] = (r["birim_fiyat"], r["kur"]) if r else (0, 0)
    recete = _recete_map(conn, bid)
    rows = conn.execute(
        """SELECT i.fatura_no, i.tarih, i.musteri, k.urun_adi, m.ad mamul_ad,
                  COALESCE(NULLIF(k.gtip,''), m.gtip) gtip, m.satir_kodu, k.miktar_kg, k.mamul_id
           FROM ihracat_kalem k JOIN ihracat i ON i.id=k.ihracat_id JOIN mamul m ON m.id=k.mamul_id
           WHERE i.belge_id=? ORDER BY i.tarih, i.id, k.kalem_no""", (bid,)).fetchall()

    satirlar, toplam_tl, toplam_kg = [], 0.0, 0.0
    for r in rows:
        sarflar, tl = [], 0.0
        for h in ham:
            kg = round(r["miktar_kg"] * recete.get(r["mamul_id"], {}).get(h["id"], 0), 2)
            sarflar.append(kg or 0)
            bf, kur = son_fiyat[h["id"]]
            if kg and bf and kur:
                tl += kg * bf * kur
        toplam_tl += tl
        toplam_kg += r["miktar_kg"]
        satirlar.append([r["fatura_no"], r["tarih"], r["musteri"], r["urun_adi"] or r["mamul_ad"],
                         r["gtip"], r["satir_kodu"], r["miktar_kg"]] + sarflar + [round(tl, 2)])
    satirlar.append(["TOPLAM", "", "", "", "", "", round(toplam_kg, 2)] + [""] * len(ham) + [round(toplam_tl, 2)])
    conn.close()
    n = len(ham)
    return {
        "baslik": "DİİB KAPSAMI KDV İSTİSNA LİSTESİ (TASLAK)",
        "altbilgi": f"Belge: {belge['belge_no']} · {firma['unvan']} · Rapor tarihi: {date.today().isoformat()}",
        "dosya": f"kdv-istisna-{belge['belge_no'].replace('/', '-')}",
        "sayfalar": [{
            "ad": "KDV İstisna Listesi",
            "basliklar": ["Fatura No", "Tarih", "Müşteri", "Ürün", "GTİP", "DİİB Satır Kodu", "İhraç kg"]
                        + [f"{h['ad']} sarf kg" for h in ham] + ["KDV'siz Düşülecek Tutar (TL)"],
            "genislikler": [18, 11, 22, 24, 17, 16, 10] + [13] * n + [18],
            "num_cols": set(range(7, 9 + n)),
            "satirlar": satirlar,
            "not": "TASLAK — Sarflar kapasite raporu reçetelerine göredir. TL tutar = sarf kg × son DİİB "
                   "ithalat birim fiyatı × ithalat tarihindeki TCMB kuru. YMM/müşavir kontrolü gerekir.",
        }],
    }


def build_tev(user):
    conn, belge, firma = _ctx(user)
    bid = belge["id"]
    konu_ham = {}
    for r in conn.execute(
        """SELECT DISTINCT h.id, h.ad, i.mense FROM ithalat_kalem k
           JOIN ithalat i ON i.id=k.ithalat_id JOIN hammadde h ON h.id=k.hammadde_id
           WHERE i.belge_id=?""", (bid,)):
        if any(u in (r["mense"] or "").upper() for u in UCUNCU_ULKE_MENSE):
            konu_ham[r["id"]] = r["ad"]
    son_fiyat = {}
    for hid in konu_ham:
        r = conn.execute(
            """SELECT k.birim_fiyat FROM ithalat_kalem k JOIN ithalat i ON i.id=k.ithalat_id
               WHERE i.belge_id=? AND k.hammadde_id=? AND k.birim_fiyat > 0
               ORDER BY i.tarih DESC LIMIT 1""", (bid, hid)).fetchone()
        son_fiyat[hid] = r["birim_fiyat"] if r else 0
    recete = _recete_map(conn, bid)

    satirlar, toplam = [], 0.0
    for r in conn.execute(
        """SELECT i.fatura_no, i.beyanname_no, i.tarih, i.ulke, k.urun_adi, m.ad mamul_ad,
                  m.satir_kodu, k.miktar_kg, k.mamul_id
           FROM ihracat_kalem k JOIN ihracat i ON i.id=k.ihracat_id JOIN mamul m ON m.id=k.mamul_id
           WHERE i.belge_id=? ORDER BY i.tarih, i.id, k.kalem_no""", (bid,)).fetchall():
        if not any(u in (r["ulke"] or "").upper() for u in TEV_ULKELERI):
            continue
        for hid, had in konu_ham.items():
            kats = recete.get(r["mamul_id"], {}).get(hid, 0)
            if not kats:
                continue
            sarf = round(r["miktar_kg"] * kats, 2)
            tev = round(sarf * son_fiyat[hid] * TEV_ORANI, 2)
            toplam += tev
            satirlar.append([r["fatura_no"], r["beyanname_no"], r["tarih"], r["ulke"],
                             r["urun_adi"] or r["mamul_ad"], r["satir_kodu"], r["miktar_kg"],
                             had, sarf, son_fiyat[hid], TEV_ORANI * 100, tev])
    satirlar.append(["TOPLAM", "", "", "", "", "", "", "", "", "", "", round(toplam, 2)])
    conn.close()
    return {
        "baslik": "EK-8 TELAFİ EDİCİ VERGİ TABLOSU (TASLAK)",
        "altbilgi": f"Belge: {belge['belge_no']} · {firma['unvan']} · Rapor tarihi: {date.today().isoformat()}",
        "dosya": f"tev-ek8-{belge['belge_no'].replace('/', '-')}",
        "sayfalar": [{
            "ad": "TEV Tablosu",
            "basliklar": ["İhracat Fatura", "Beyanname", "Tarih", "Varış Ülkesi", "Ürün", "Satır Kodu",
                          "İhraç kg", "TEV'e Konu Hammadde", "Sarf kg", "Birim Fiyat $", "Oran %", "TEV $"],
            "genislikler": [17, 19, 11, 14, 22, 15, 10, 20, 10, 11, 8, 12],
            "num_cols": {7, 9, 10, 11, 12},
            "satirlar": satirlar,
            "not": f"TASLAK — AB/STA varışlı ihracatta 3. ülke ({', '.join(UCUNCU_ULKE_MENSE)}) menşeli girdiler "
                   f"için %{TEV_ORANI*100:.1f} ile hesaplanmıştır. İç piyasadan karşılanan girdiler TEV doğurmaz — "
                   "müşavir kontrolü ve beyanname bazlı eşleştirme onayı gerekir.",
        }],
    }


def build_kapatma(user):
    conn, belge, firma = _ctx(user)
    bid = belge["id"]
    alt = f"Belge: {belge['belge_no']} · {firma['unvan']} · Rapor tarihi: {date.today().isoformat()}"
    sayfalar = []

    s1 = []
    for m in conn.execute("SELECT * FROM mamul WHERE belge_id=? ORDER BY satir_kodu", (bid,)):
        g = conn.execute(
            """SELECT COALESCE(SUM(k.miktar_kg),0) kg, COALESCE(SUM(k.tutar),0) t
               FROM ihracat_kalem k JOIN ihracat i ON i.id=k.ihracat_id
               WHERE i.belge_id=? AND k.mamul_id=?""", (bid, m["id"])).fetchone()
        oran = round(100 * g["kg"] / m["taahhut_kg"], 2) if m["taahhut_kg"] else 0
        s1.append([m["satir_kodu"], m["gtip"], m["ad"], m["taahhut_kg"], g["kg"],
                   m["taahhut_kg"] - g["kg"], oran, round(g["t"], 2)])
    sayfalar.append({"ad": "Taahhüt Gerçekleşme",
                     "basliklar": ["Satır Kodu", "GTİP", "Mamul", "Taahhüt kg", "Gerçekleşen kg",
                                   "Kalan kg", "Gerçekleşme %", "Gerçekleşen Tutar"],
                     "genislikler": [15, 17, 40, 12, 13, 12, 12, 14],
                     "num_cols": {4, 5, 6, 7, 8}, "satirlar": s1, "not": ""})

    s2 = [[r["beyanname_no"], r["tarih"], r["gumruk"], r["satici"], r["mense"], r["kalem_no"],
           r["gtip"], r["aciklama"] or r["ad"], r["miktar_kg"], r["birim_fiyat"], r["tutar"],
           r["doviz"], r["kur"]]
          for r in conn.execute(
        """SELECT i.beyanname_no, i.tarih, i.gumruk, i.satici, i.mense, i.doviz, i.kur,
                  k.kalem_no, COALESCE(NULLIF(k.gtip,''), h.gtip) gtip, k.aciklama,
                  h.ad, k.miktar_kg, k.birim_fiyat, k.tutar
           FROM ithalat_kalem k JOIN ithalat i ON i.id=k.ithalat_id
           JOIN hammadde h ON h.id=k.hammadde_id
           WHERE i.belge_id=? ORDER BY i.tarih, i.id, k.kalem_no""", (bid,))]
    sayfalar.append({"ad": "İthalat Listesi",
                     "basliklar": ["Beyanname No", "Tarih", "Gümrük", "Satıcı", "Menşe", "Kalem No",
                                   "GTİP", "Madde", "Miktar kg", "Birim Fiyat", "Tutar", "Döviz", "Kur"],
                     "genislikler": [20, 11, 22, 26, 10, 9, 17, 24, 11, 11, 12, 8, 9],
                     "num_cols": {9, 10, 11, 13}, "satirlar": s2, "not": ""})

    s3 = [[r["fatura_no"], r["beyanname_no"], r["tarih"], r["musteri"], r["ulke"], r["kalem_no"],
           r["gtip"], r["urun_adi"] or r["mamul_ad"], r["satir_kodu"], r["miktar_kg"],
           r["tutar"], r["doviz"], r["kapanis_tarihi"] or "AÇIK"]
          for r in conn.execute(
        """SELECT i.fatura_no, i.beyanname_no, i.tarih, i.musteri, i.ulke, i.doviz,
                  i.kapanis_tarihi, k.kalem_no, COALESCE(NULLIF(k.gtip,''), m.gtip) gtip,
                  k.urun_adi, m.ad mamul_ad, m.satir_kodu, k.miktar_kg, k.tutar
           FROM ihracat_kalem k JOIN ihracat i ON i.id=k.ihracat_id JOIN mamul m ON m.id=k.mamul_id
           WHERE i.belge_id=? ORDER BY i.tarih, i.id, k.kalem_no""", (bid,))]
    sayfalar.append({"ad": "İhracat Listesi",
                     "basliklar": ["Fatura No", "Beyanname No", "Tarih", "Müşteri", "Ülke", "Kalem No",
                                   "GTİP", "Ürün", "Satır Kodu", "Miktar kg", "Tutar", "Döviz", "Kapanış"],
                     "genislikler": [18, 20, 11, 26, 12, 9, 17, 26, 15, 11, 12, 8, 12],
                     "num_cols": {10, 11}, "satirlar": s3, "not": ""})

    s4 = []
    for h in conn.execute("SELECT * FROM hammadde WHERE firma_id=? ORDER BY yerli, ad", (user["firma_id"],)):
        i = conn.execute(
            """SELECT COALESCE(SUM(k.miktar_kg),0) s FROM ithalat_kalem k
               JOIN ithalat it ON it.id=k.ithalat_id WHERE it.belge_id=? AND k.hammadde_id=?""",
            (bid, h["id"])).fetchone()["s"]
        s = conn.execute(
            """SELECT COALESCE(SUM(k.miktar_kg * r.katsayi),0) s FROM ihracat_kalem k
               JOIN ihracat ih ON ih.id=k.ihracat_id JOIN recete r ON r.mamul_id=k.mamul_id
               WHERE ih.belge_id=? AND r.hammadde_id=?""", (bid, h["id"])).fetchone()["s"]
        if not i and not s:
            continue
        s4.append([h["satir_kodu"], h["gtip"], h["ad"], round(i, 1), round(s, 1), round(i - s, 1),
                   "Eşdeğer eşya (iç piyasa) kullanımı" if i - s < 0 else ""])
    sayfalar.append({"ad": "Hammadde Sarf Özeti",
                     "basliklar": ["Satır Kodu", "GTİP", "Hammadde", "İthal Edilen kg",
                                   "Reçete Sarfı kg", "Bakiye kg", "Not"],
                     "genislikler": [15, 17, 24, 14, 14, 12, 30],
                     "num_cols": {4, 5, 6}, "satirlar": s4,
                     "not": "TASLAK — Kapatma başvurusu öncesi müşavir/YMM kontrolü gerekir. "
                            "Eksi bakiye: eşdeğer eşya kullanımı."})
    conn.close()
    return {"baslik": "DİİB KAPATMA DOSYASI SETİ (TASLAK)", "altbilgi": alt,
            "dosya": f"kapatma-seti-{belge['belge_no'].replace('/', '-')}", "sayfalar": sayfalar}


# ================================================================ XLSX
_thin = Side(style="thin", color="B7C4D0")
BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def render_xlsx(rapor):
    wb = Workbook()
    wb.remove(wb.active)
    for sf in rapor["sayfalar"]:
        ws = wb.create_sheet(sf["ad"][:31])
        ws.append([rapor["baslik"] + (f" — {sf['ad']}" if len(rapor["sayfalar"]) > 1 else "")])
        ws["A1"].font = Font(bold=True, size=13, color="0F4C81")
        ws.append([rapor["altbilgi"]])
        ws.append([])
        ws.append(sf["basliklar"])
        hr = ws.max_row
        for i in range(1, len(sf["basliklar"]) + 1):
            c = ws.cell(row=hr, column=i)
            c.fill = PatternFill("solid", fgColor="0F4C81")
            c.font = Font(bold=True, color="FFFFFF", size=10)
            c.border = BORDER
            c.alignment = Alignment(horizontal="center", wrap_text=True)
            g = sf.get("genislikler")
            if g and i <= len(g) and g[i - 1]:
                ws.column_dimensions[get_column_letter(i)].width = g[i - 1]
        for row in sf["satirlar"]:
            ws.append(row)
            r = ws.max_row
            for i in range(1, len(row) + 1):
                c = ws.cell(row=r, column=i)
                c.border = BORDER
                c.font = Font(size=10, bold=(row and row[0] == "TOPLAM"))
                if i in sf.get("num_cols", set()):
                    c.number_format = "#,##0.00"
                    c.alignment = Alignment(horizontal="right")
        if sf.get("not"):
            ws.append([])
            ws.append([sf["not"]])
            ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9, color="666666")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{rapor["dosya"]}.xlsx"'})


# ================================================================ PDF
_FONT_PATHS = [
    r"C:\Windows\Fonts\arial.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
]
_FONT_BOLD = [
    r"C:\Windows\Fonts\arialbd.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf",
]


def render_pdf(rapor):
    from fpdf import FPDF
    font = next((p for p in _FONT_PATHS if os.path.exists(p)), None)
    fontb = next((p for p in _FONT_BOLD if os.path.exists(p)), None)
    if not font:
        raise HTTPException(500, "PDF için TTF font bulunamadı (DejaVu/Arial)")

    pdf = FPDF(orientation="L", format="A4")
    pdf.set_auto_page_break(True, margin=12)
    pdf.add_font("TR", "", font)
    pdf.add_font("TR", "B", fontb or font)

    for sf in rapor["sayfalar"]:
        pdf.add_page()
        pdf.set_font("TR", "B", 13)
        pdf.set_text_color(15, 76, 129)
        pdf.cell(0, 8, rapor["baslik"] + (f" — {sf['ad']}" if len(rapor["sayfalar"]) > 1 else ""),
                 new_x="LMARGIN", new_y="NEXT")
        pdf.set_font("TR", "", 8)
        pdf.set_text_color(90, 90, 90)
        pdf.cell(0, 5, rapor["altbilgi"], new_x="LMARGIN", new_y="NEXT")
        pdf.ln(2)

        n = len(sf["basliklar"])
        genis = sf.get("genislikler") or [1] * n
        toplam_g = sum(genis[:n]) or n
        col_w = [max(14, g / toplam_g * (pdf.w - 24)) for g in genis[:n]]
        # normalize (toplam sayfa genişliğine sığdır)
        k = (pdf.w - 24) / sum(col_w)
        col_w = [w * k for w in col_w]

        pdf.set_text_color(0, 0, 0)
        with pdf.table(col_widths=col_w, text_align="LEFT", line_height=4.6,
                       borders_layout="ALL", first_row_as_headings=True,
                       padding=1) as table:
            hrow = table.row()
            pdf.set_font("TR", "B", 6.8)
            for hcell in sf["basliklar"]:
                hrow.cell(str(hcell))
            pdf.set_font("TR", "", 6.6)
            for row in sf["satirlar"]:
                trow = table.row()
                for i, v in enumerate(row, 1):
                    if isinstance(v, float):
                        v = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                    trow.cell("" if v is None else str(v))
        if sf.get("not"):
            pdf.ln(2)
            pdf.set_font("TR", "", 7)
            pdf.set_text_color(110, 110, 110)
            pdf.multi_cell(0, 4, sf["not"])

    out = pdf.output()
    return StreamingResponse(
        io.BytesIO(bytes(out)), media_type="application/pdf",
        headers={"Content-Disposition": f'inline; filename="{rapor["dosya"]}.pdf"'})


# ================================================================ HTML (görüntüleme)
def render_html(rapor):
    e = html_mod.escape
    parts = [f"""<!DOCTYPE html><html lang="tr"><head><meta charset="utf-8">
<title>{e(rapor['baslik'])}</title>
<style>
body{{font-family:-apple-system,'Segoe UI',Roboto,sans-serif;margin:24px;color:#1c2733;background:#fff}}
h1{{font-size:19px;color:#0f4c81;margin-bottom:2px}}
p.alt{{color:#66788c;font-size:12px;margin:0 0 18px}}
h2{{font-size:15px;color:#0f4c81;margin:26px 0 8px}}
table{{border-collapse:collapse;width:100%;font-size:12px;margin-bottom:6px}}
th{{background:#0f4c81;color:#fff;padding:6px 7px;text-align:left;font-size:10.5px;position:sticky;top:0}}
td{{border:1px solid #d5dfe9;padding:5px 7px;white-space:nowrap}}
tr:nth-child(even) td{{background:#f7fafc}}
tr.toplam td{{font-weight:700;background:#eef4fb}}
td.num{{text-align:right;font-variant-numeric:tabular-nums}}
p.not{{color:#8a7430;background:#fdf6e8;border:1px solid #f2e2bd;border-radius:8px;padding:9px 12px;font-size:11.5px}}
.bar{{position:sticky;top:0;background:#fff;padding:8px 0;border-bottom:1px solid #e3e9f0;margin-bottom:14px;display:flex;gap:10px;align-items:center;z-index:5}}
.bar a{{text-decoration:none;background:#1273c4;color:#fff;padding:8px 16px;border-radius:8px;font-size:13px;font-weight:600}}
.bar a.g{{background:#fff;color:#1c2733;border:1px solid #d5dfe9}}
@media print{{.bar{{display:none}}td,th{{font-size:9px;padding:3px 4px}}}}
.twrap{{overflow-x:auto}}
</style></head><body>
<div class="bar">
  <a href="?format=xlsx">⬇ Excel indir</a>
  <a href="?format=pdf" target="_blank">⬇ PDF indir</a>
  <a class="g" href="javascript:print()">🖨 Yazdır</a>
</div>
<h1>{e(rapor['baslik'])}</h1><p class="alt">{e(rapor['altbilgi'])}</p>"""]
    for sf in rapor["sayfalar"]:
        if len(rapor["sayfalar"]) > 1:
            parts.append(f"<h2>{e(sf['ad'])}</h2>")
        parts.append("<div class='twrap'><table><thead><tr>"
                     + "".join(f"<th>{e(str(h))}</th>" for h in sf["basliklar"]) + "</tr></thead><tbody>")
        for row in sf["satirlar"]:
            cls = ' class="toplam"' if row and row[0] == "TOPLAM" else ""
            cells = []
            for i, v in enumerate(row, 1):
                if isinstance(v, float):
                    txt = f"{v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                else:
                    txt = "" if v is None else str(v)
                nc = ' class="num"' if i in sf.get("num_cols", set()) else ""
                cells.append(f"<td{nc}>{e(txt)}</td>")
            parts.append(f"<tr{cls}>{''.join(cells)}</tr>")
        parts.append("</tbody></table></div>")
        if sf.get("not"):
            parts.append(f"<p class='not'>{e(sf['not'])}</p>")
    parts.append("</body></html>")
    return HTMLResponse("".join(parts))


# ================================================================ UÇLAR
_BUILDERS = {"kdv": build_kdv, "tev": build_tev, "kapatma": build_kapatma}


@router.get("/{tip}")
def rapor(tip: str, format: str = "xlsx", user=Depends(auth.require_user)):
    if tip not in _BUILDERS:
        raise HTTPException(404, "Rapor: kdv | tev | kapatma")
    rapor_data = _BUILDERS[tip](user)
    if format == "pdf":
        return render_pdf(rapor_data)
    if format == "html":
        return render_html(rapor_data)
    return render_xlsx(rapor_data)
